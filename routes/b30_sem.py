from flask import Blueprint, request, jsonify, session, render_template, current_app, send_file
from routes.shared import cruc_client, publish_barcode
from crucible import Dataset
from crucible.utils import get_tz_isoformat
from config import B30_SEM_CONFIG
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from threading import Lock
import requests
import io
import csv
import os
import json
import tempfile
import tifffile

b30_sem_bp = Blueprint("b30_sem", __name__)

PRINTER_NAME = B30_SEM_CONFIG.get("printer_name", "crucible-printer/b30-122")
LA_TZ = ZoneInfo("America/Los_Angeles")

_ORCID_NAME_CACHE = {}
_ORCID_NAME_CACHE_LOCK = Lock()

_INCREMENTAL_CACHE = {}
_INCREMENTAL_CACHE_LOCK = Lock()


# ---------- Helpers ----------

def _parse_ts(ts):
    if not ts:
        return datetime.min.replace(tzinfo=timezone.utc)
    try:
        dt = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return datetime.min.replace(tzinfo=timezone.utc)


def _pick(d, *keys, default=""):
    for k in keys:
        v = d.get(k) if isinstance(d, dict) else None
        if v is not None and str(v).strip() != "":
            return v
    return default


def _clean(x):
    return "" if x is None else str(x).strip()


def _get_state():
    if "b30_sem" not in session:
        session["b30_sem"] = {
            "sample_unique_id": "",
            "sample_name": "",
            "sample_type": "",
            "sample_description": "",
        }
    return session["b30_sem"]


def _dataset_to_row(details):
    sci = details.get("scientific_metadata") or {}
    dt = _parse_ts(details.get("timestamp"))
    date_str = dt.astimezone(LA_TZ).strftime("%Y-%m-%d %H:%M") if dt else ""

    owner_orcid = _pick(details, "owner_orcid", default="")
    user_name = get_name_from_orcid_cached(owner_orcid) if owner_orcid else ""

    return {
        "Date": date_str,
        "User": user_name or owner_orcid,
        "Vacuum": _clean(sci.get("vacuum_level")),
        "Spot": _clean(sci.get("spot_size")),
        "HV (V)": _clean(sci.get("high_voltage_V")),
        "Current (A)": _clean(sci.get("emission_current_A")),
        "Pressure (Torr)": _clean(sci.get("chamber_pressure_Torr")),
        "EDX": "Yes" if sci.get("edx_used") in (True, "true", "True", 1) else "",
        "Energy (keV)": _clean(sci.get("primary_energy_keV")),
        "Comment": _clean(sci.get("comment")),
        "_dataset_id": details.get("unique_id") or details.get("dsid") or "",
        "_timestamp": details.get("timestamp") or "",
    }


def get_name_from_orcid(orcid_id):
    url = f"https://orcid.org/{orcid_id}"
    headers = {"Accept": "application/json"}
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        name_data = data.get("person", {}).get("name", {})
        given_names = ""
        family_name = ""
        credit_name = ""
        if "given-names" in name_data:
            given_names = name_data.get("given-names", {}).get("value", "")
        if "family-name" in name_data:
            family_name = name_data.get("family-name", {}).get("value", "")
        if "credit-name" in name_data:
            credit_name = ((name_data or {}).get("credit-name") or {}).get("value", "")
        if credit_name:
            return credit_name
        elif given_names or family_name:
            return f"{given_names} {family_name}".strip()
        else:
            return "Name is private or not set."
    except requests.exceptions.RequestException as e:
        return f"Error fetching data: {e}"


def get_name_from_orcid_cached(orcid_id: str) -> str:
    oid = (orcid_id or "").strip()
    if not oid:
        return ""
    with _ORCID_NAME_CACHE_LOCK:
        if oid in _ORCID_NAME_CACHE:
            return _ORCID_NAME_CACHE[oid]
    name = get_name_from_orcid(oid)
    if not name:
        name = oid
    with _ORCID_NAME_CACHE_LOCK:
        _ORCID_NAME_CACHE[oid] = name
    return name


def _get_cache_bucket(key):
    with _INCREMENTAL_CACHE_LOCK:
        if key not in _INCREMENTAL_CACHE:
            _INCREMENTAL_CACHE[key] = {
                "row_by_id": {},
                "ts_by_id": {},
                "ordered_ids": [],
                "last_scan_at": None,
            }
        return _INCREMENTAL_CACHE[key]


def _incremental_refresh_rows(project_id, instrument_name):
    key = (project_id, instrument_name)
    bucket = _get_cache_bucket(key)

    all_ds = cruc_client.datasets.list(
        project_id=project_id,
        instrument_name=instrument_name,
        limit=2000
    )
    summaries = sorted(all_ds, key=lambda d: _parse_ts(d.get("timestamp")), reverse=True)
    summary_by_id = {d["unique_id"]: d for d in summaries if d.get("unique_id")}
    current_ids = set(summary_by_id.keys())

    row_by_id = bucket["row_by_id"]
    ts_by_id = bucket["ts_by_id"]

    for dsid in list(row_by_id.keys()):
        if dsid not in current_ids:
            row_by_id.pop(dsid, None)
            ts_by_id.pop(dsid, None)

    for dsid, s in summary_by_id.items():
        s_ts = _parse_ts(s.get("timestamp"))
        cached_ts = ts_by_id.get(dsid)
        need_fetch = (dsid not in row_by_id) or (cached_ts is None) or (s_ts > cached_ts)
        if need_fetch:
            details = cruc_client.datasets.get(dsid=dsid, include_metadata=True)
            row_by_id[dsid] = _dataset_to_row(details)
            ts_by_id[dsid] = _parse_ts(details.get("timestamp"))

    ordered_ids = sorted(ts_by_id.keys(), key=lambda i: ts_by_id[i], reverse=True)
    bucket["ordered_ids"] = ordered_ids
    bucket["last_scan_at"] = datetime.now(timezone.utc)

    return [row_by_id[i] for i in ordered_ids]


# ---------- TIFF metadata extraction ----------

def _find_fei_value(metadata: dict, *pairs):
    """Returns the first non-empty value found at any of the given (section, key) pairs."""
    for section, key in pairs:
        value = (metadata.get(section) or {}).get(key)
        if value is not None and str(value).strip() != "":
            return value
    return None


def _parse_fei_metadata(metadata: dict) -> dict:
    """
    Maps an FEI SEM metadata dict (as returned by tifffile's `fei_metadata`
    property, e.g. {"Vacuum": {"UserMode": ..., "ChPressure": ...}, "EBeam": {...}})
    to our config keys.
    """
    result = {}

    vacuum_raw = _find_fei_value(metadata, ("Vacuum", "UserMode"))
    if vacuum_raw:
        v = str(vacuum_raw).lower()
        if "esem" in v or "wet" in v or "environ" in v:
            result["vacuum_level"] = "ESEM"
        elif "low" in v:
            result["vacuum_level"] = "Low"
        else:
            result["vacuum_level"] = "High"

    spot = _find_fei_value(metadata, ("Beam", "Spot"), ("EBeam", "Spot"))
    if spot is not None:
        result["spot_size"] = str(spot)

    # High voltage — stored in V, may need conversion
    hv = _find_fei_value(metadata, ("Beam", "HV"), ("EBeam", "HV"))
    if hv is not None:
        result["high_voltage_V"] = str(hv)

    current = _find_fei_value(metadata, ("EBeam", "EmissionCurrent"))
    if current is not None:
        result["emission_current_A"] = str(current)

    pressure = _find_fei_value(metadata, ("Vacuum", "ChPressure"))
    if pressure is not None:
        result["chamber_pressure_Torr"] = str(pressure)

    return result


def _sanitize_metadata_value(value):
    """Recursively coerce a value into something JSON-serializable."""
    if isinstance(value, dict):
        return {str(k): _sanitize_metadata_value(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_sanitize_metadata_value(v) for v in value]
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _extract_full_tiff_metadata(fobj):
    """Read the FEI metadata for a single uploaded TIFF FileStorage, leaving its
    stream position reset to 0 so it can still be saved to disk afterwards."""
    try:
        file_bytes = fobj.read()
        fobj.seek(0)
        with tifffile.TiffFile(io.BytesIO(file_bytes)) as tif:
            metadata = tif.fei_metadata
    except Exception:
        return None
    if not metadata:
        return None
    return _sanitize_metadata_value(metadata)


# ---------- Routes ----------

@b30_sem_bp.route("/")
def page():
    return render_template("b30_sem.html", config=B30_SEM_CONFIG)


@b30_sem_bp.route("/api/state", methods=["GET"])
def get_state():
    return jsonify(_get_state())


# ---------- TIFF metadata ----------

@b30_sem_bp.route("/api/extract-tiff-metadata", methods=["POST"])
def extract_tiff_metadata():
    """Receive a TIFF file, extract FEI SEM metadata, return mapped field values."""
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400

    filename = f.filename or ""
    if not filename.lower().endswith((".tif", ".tiff")):
        return jsonify({"error": "File must be a TIFF (.tif or .tiff)"}), 400

    try:
        with tifffile.TiffFile(io.BytesIO(f.read())) as tif:
            metadata = tif.fei_metadata
    except Exception as e:
        current_app.logger.warning(f"[b30_sem] tifffile parse error: {e}")
        return jsonify({"found": False, "fields": {}, "message": "Could not read this TIFF file. Please enter values manually."}), 200

    if not metadata:
        return jsonify({"found": False, "fields": {}, "message": "No FEI metadata found in this TIFF. Please enter values manually."}), 200

    fields = _parse_fei_metadata(metadata)
    return jsonify({"found": bool(fields), "fields": fields}), 200


# ---------- EDX spectrum ----------

@b30_sem_bp.route("/api/parse-edx-spectrum", methods=["POST"])
def parse_edx_spectrum():
    """
    Receive an EDX spectrum XLSX file, extract metadata from the header rows,
    and return metadata fields + spectrum data as CSV.
    """
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file provided"}), 400

    filename = f.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "File must be an Excel file (.xlsx or .xls)"}), 400

    file_bytes = f.read()

    try:
        import openpyxl
    except ImportError:
        return jsonify({"error": "openpyxl is not installed. Run: pip install openpyxl"}), 500

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        metadata = {}
        spectrum_start_row = None

        # Scan header rows for known metadata keys
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row is None or all(v is None for v in row):
                continue

            # Look for key-value pairs in first two columns
            key_cell = str(row[0] or "").strip()
            val_cell = row[1] if len(row) > 1 else None

            key_lower = key_cell.lower()

            if any(k in key_lower for k in ("kv", "acc", "voltage", "hv", "energy")):
                if val_cell is not None:
                    try:
                        metadata["primary_energy_keV"] = str(float(val_cell))
                    except (TypeError, ValueError):
                        metadata["primary_energy_keV"] = str(val_cell)

            # Detect start of spectrum data (numeric keV/counts columns)
            if key_lower in ("kev", "energy (kev)", "channel") or (
                val_cell is not None and isinstance(val_cell, (int, float)) and isinstance(key_cell, str) and key_lower.replace('.', '', 1).isdigit()
            ):
                spectrum_start_row = row_idx
                break

        # Build CSV of spectrum data from spectrum_start_row onwards
        sio = io.StringIO()
        writer = csv.writer(sio)
        if spectrum_start_row is not None:
            for row in ws.iter_rows(min_row=spectrum_start_row, values_only=True):
                if row and any(v is not None for v in row):
                    writer.writerow([v if v is not None else "" for v in row])
        spectrum_csv = sio.getvalue()

        return jsonify({
            "metadata": metadata,
            "spectrum_csv": spectrum_csv,
        }), 200

    except Exception as e:
        current_app.logger.error(f"[b30_sem] EDX parse error: {e}")
        return jsonify({"error": f"Failed to parse file: {e}"}), 500


# ---------- Sample lookup (barcode scan) ----------

@b30_sem_bp.route("/api/lookup-sample", methods=["POST"])
def lookup_sample():
    data = request.get_json()
    unique_id = data.get("unique_id", "").strip()
    if not unique_id:
        return jsonify({"error": "No barcode value provided"}), 400

    try:
        sample = cruc_client.samples.get(unique_id)
    except Exception:
        return jsonify({"found": False, "unique_id": unique_id})

    if sample is None:
        return jsonify({"found": False, "unique_id": unique_id})

    state = _get_state()
    state["sample_unique_id"] = sample["unique_id"]
    state["sample_name"] = sample["sample_name"]
    state["sample_type"] = sample.get("sample_type", "")
    state["sample_description"] = sample.get("description", "")
    session.modified = True

    return jsonify({
        "found": True,
        "unique_id": sample["unique_id"],
        "sample_name": sample["sample_name"],
        "sample_type": sample.get("sample_type", ""),
        "description": sample.get("description", ""),
    })


# ---------- Sample creation ----------

@b30_sem_bp.route("/api/create-sample", methods=["POST"])
def create_sample():
    user = session.get("user")
    if not user:
        return jsonify({"error": "Not logged in"}), 401

    data = request.get_json()
    sample_name = data.get("sample_name", "").strip()
    sample_type = data.get("sample_type", "").strip()
    description = data.get("description", "").strip()

    if not sample_name or not sample_type:
        return jsonify({"error": "sample_name and sample_type are required"}), 400

    try:
        returned_sample = cruc_client.samples.create(
            sample_name=sample_name,
            timestamp=get_tz_isoformat(),
            owner_orcid=user["orcid"],
            project_id=user["selected_project"],
            sample_type=sample_type,
            description=description or None,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    state = _get_state()
    state["sample_unique_id"] = returned_sample["unique_id"]
    state["sample_name"] = returned_sample["sample_name"]
    state["sample_type"] = sample_type
    state["sample_description"] = description
    session.modified = True

    return jsonify({
        "unique_id": returned_sample["unique_id"],
        "sample_name": returned_sample["sample_name"],
        "sample_type": sample_type,
        "description": description,
    })


# ---------- Barcode printing ----------

@b30_sem_bp.route("/api/print-barcode", methods=["POST"])
def print_barcode():
    data = request.get_json() or {}
    sample_name = data.get("sample_name", "").strip()
    sample_mfid = data.get("sample_id", "").strip()

    if not sample_mfid:
        return jsonify({"error": "sample_id is required"}), 400

    try:
        publish_barcode(PRINTER_NAME, sample_mfid, sample_name)
    except Exception as e:
        current_app.logger.error(f"[b30_sem] Barcode print failed: {e}")
        return jsonify({"error": str(e)}), 500

    return jsonify({"ok": True, "sample_id": sample_mfid, "sample_name": sample_name}), 200


# ---------- Dataset upload ----------

@b30_sem_bp.route("/api/upload-dataset", methods=["POST"])
def upload_dataset():
    """
    Create a SEM dataset in Crucible.
    Accepts multipart/form-data: metadata fields as JSON in 'metadata' part,
    optional SEM image files in 'sem_images', optional EDX images in 'edx_images',
    optional EDX spectrum CSV in 'edx_spectrum'.
    """
    user = session.get("user")
    if not user:
        return jsonify({"error": "Not logged in"}), 401

    state = _get_state()
    if not state.get("sample_unique_id"):
        return jsonify({"error": "No sample selected. Scan a barcode first."}), 400

    # Support both multipart (with files) and JSON-only requests
    if request.content_type and "multipart/form-data" in request.content_type:
        raw_meta = request.form.get("metadata", "{}")
        try:
            data = json.loads(raw_meta)
        except json.JSONDecodeError:
            return jsonify({"error": "Invalid metadata JSON"}), 400
        sem_images = request.files.getlist("sem_images")
        edx_images = request.files.getlist("edx_images")
        edx_spectrum = request.files.get("edx_spectrum")
    else:
        data = request.get_json() or {}
        sem_images = []
        edx_images = []
        edx_spectrum = None

    # Build scientific metadata from config-defined fields
    scientific_metadata = {}
    for field in B30_SEM_CONFIG["dataset_fields"]:
        key = field["key"]
        value = data.get(key)
        if value is None:
            continue
        if isinstance(value, str):
            value = value.strip()
        if value != "" and value is not None:
            scientific_metadata[key] = value

    # Capture the *full* FEI metadata embedded in any uploaded TIFFs, not just
    # the handful of fields mapped onto the form.
    tiff_metadata_by_file = {}
    for fobj in sem_images:
        if not fobj or not fobj.filename:
            continue
        if os.path.splitext(fobj.filename)[1].lower() not in (".tif", ".tiff"):
            continue
        full_meta = _extract_full_tiff_metadata(fobj)
        if full_meta:
            tiff_metadata_by_file[fobj.filename] = full_meta
    if tiff_metadata_by_file:
        scientific_metadata["tiff_metadata"] = tiff_metadata_by_file

    sample_name = (state.get("sample_name") or "").strip() or "unknown-sample"
    date_str = datetime.now(LA_TZ).strftime("%Y%m%d_%H%M%S")
    dataset_name = f"{date_str}_SEM_on_{sample_name}"

    try:
        ds = Dataset(
            dataset_name=dataset_name,
            dataset_type=B30_SEM_CONFIG["dataset_type"],
            owner_orcid=user["orcid"],
            project_id=user["selected_project"],
            instrument_name=B30_SEM_CONFIG["instrument_name"],
            measurement=B30_SEM_CONFIG["measurement"],
            timestamp=get_tz_isoformat(),
        )
        new_dataset = cruc_client.datasets.create(ds, scientific_metadata=scientific_metadata)
        dataset_id = new_dataset["dsid"]

        cruc_client.datasets.add_sample(
            dataset_id=dataset_id,
            sample_id=state["sample_unique_id"],
        )

        # Upload files if provided
        uploaded_files = []
        all_file_objs = (
            [("sem", f) for f in sem_images if f and f.filename]
            + [("edx_img", f) for f in edx_images if f and f.filename]
            + ([("edx_spectrum", edx_spectrum)] if edx_spectrum and edx_spectrum.filename else [])
        )

        for _kind, fobj in all_file_objs:
            # Write to a temp file so cruc_client.files can read it from disk
            suffix = os.path.splitext(fobj.filename)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                fobj.save(tmp)
                tmp_path = tmp.name
            try:
                cruc_client.files.add_file_to_dataset(dataset_id, tmp_path)
                uploaded_files.append(fobj.filename)
            except Exception as e:
                current_app.logger.warning(f"[b30_sem] File upload failed for {fobj.filename}: {e}")
            finally:
                os.unlink(tmp_path)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({
        "dataset_name": dataset_name,
        "dataset_id": dataset_id,
        "sample_name": state["sample_name"],
        "uploaded_files": uploaded_files,
    })


# ---------- Dataset logbook ----------

@b30_sem_bp.route("/api/recent-datasets", methods=["GET"])
def recent_datasets():
    user = session.get("user")
    if not user:
        return jsonify({"error": "Not logged in"}), 401

    project_id = user.get("selected_project")
    if not project_id:
        return jsonify({"error": "No selected project"}), 400

    instrument_name = B30_SEM_CONFIG["instrument_name"]
    limit = int(request.args.get("limit", 100))
    limit = max(1, min(limit, 5000))

    try:
        rows_all = _incremental_refresh_rows(project_id, instrument_name)
        rows = []
        for row in rows_all:
            out = dict(row)
            out.pop("_dataset_id", None)
            out.pop("_timestamp", None)
            rows.append(out)
            if len(rows) >= limit:
                break
        return jsonify({"rows": rows, "count": len(rows)}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@b30_sem_bp.route("/api/recent-datasets/b30_sem_recent_datasets.csv", methods=["GET"])
def export_recent_datasets_csv():
    user = session.get("user")
    if not user:
        return jsonify({"error": "Not logged in"}), 401

    project_id = user.get("selected_project")
    if not project_id:
        return jsonify({"error": "No selected project"}), 400

    instrument_name = B30_SEM_CONFIG["instrument_name"]
    limit = int(request.args.get("limit", 100))
    limit = max(1, min(limit, 5000))

    cols = ["Date", "User", "Vacuum", "Spot", "HV (V)", "Current (A)",
            "Pressure (Torr)", "EDX", "Energy (keV)", "Comment"]

    try:
        rows_all = _incremental_refresh_rows(project_id, instrument_name)
        selected = rows_all[:limit]

        sio = io.StringIO()
        writer = csv.DictWriter(sio, fieldnames=cols, extrasaction="ignore")
        writer.writeheader()
        for row in selected:
            writer.writerow(row)

        bio = io.BytesIO(sio.getvalue().encode("utf-8-sig"))
        bio.seek(0)

        ts = datetime.now(LA_TZ).strftime("%Y%m%d_%H%M%S")
        return send_file(
            bio,
            as_attachment=True,
            download_name=f"{ts}_sem_recent_datasets.csv",
            mimetype="text/csv",
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500
